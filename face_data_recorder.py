import cv2
import mediapipe as mp
import os
import sys
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ================= CONFIG =================
OUTPUT_DIR = "recorded_faces"
ATTENDANCE_XLSX = "attendance.xlsx"
ATTENDANCE_CSV = "attendance.csv"
TARGET_IMAGE_COUNT = 30  # Number of images to capture for training

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ================= ATTENDANCE =================
def save_attendance(name, is_present=True):
    """Saves attendance record to both Excel and CSV files."""
    today = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    # --- EXCEL ---
    try:
        if os.path.exists(ATTENDANCE_XLSX):
            wb = load_workbook(ATTENDANCE_XLSX)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(today)
            ws.append(["Name", "Time"])

        # Check if attendance is already taken for today
        names_in_sheet = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        if name not in names_in_sheet:
            if is_present:
                ws.append([name, time_now])
                print(f"✅ Excel Attendance Taken: {name} at {time_now}")

        wb.save(ATTENDANCE_XLSX)
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")

    # --- CSV ---
    try:
        file_exists = os.path.exists(ATTENDANCE_CSV)
        with open(ATTENDANCE_CSV, mode='a', newline='') as file:
            writer = csv.writer(file)
            if not file_exists or os.stat(ATTENDANCE_CSV).st_size == 0:
                writer.writerow(["Date", "Name", "Time"])

            # Simple check for CSV
            writer.writerow([today, name, time_now])
    except Exception as e:
        print(f"❌ Error saving to CSV: {e}")


# ================= MAIN DATA RECORDER LOGIC =================

def record_face_data(student_name):
    """
    Captures multiple face images for a given student name
    and saves them in a structured folder (OUTPUT_DIR/Name/Image_ID.jpg).
    """

    # Use underscores for folder safety
    name = student_name.strip().replace(" ", "_").replace("/", "_").replace("\\", "_")

    if name == "Unknown" or not name:
        print("Error: Invalid student name provided.")
        return

    # Create subdirectory for the student
    student_dir = os.path.join(OUTPUT_DIR, name)
    os.makedirs(student_dir, exist_ok=True)

    mp_face_detection = mp.solutions.face_detection
    cam = cv2.VideoCapture(0)

    if not cam.isOpened():
        print("Error: Could not open camera.")
        return

    print(f"\n--- Starting Data Capture for: {student_name} ---")

    with mp_face_detection.FaceDetection(
            model_selection=0,
            min_detection_confidence=0.6
    ) as face_detection:

        img_count = 0

        while True:
            ret, frame = cam.read()
            if not ret:
                break

            frame = cv2.flip(frame, 1)  # Mirror image for better interaction
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = face_detection.process(rgb)

            if results.detections:
                for det in results.detections:
                    bbox = det.location_data.relative_bounding_box
                    h, w, _ = frame.shape

                    x = int(bbox.xmin * w)
                    y = int(bbox.ymin * h)
                    bw = int(bbox.width * w)
                    bh = int(bbox.height * h)

                    # Draw rectangle on face
                    cv2.rectangle(frame, (x, y), (x + bw, y + bh), (0, 255, 0), 2)

                    # Crop face (add a small margin for better training)
                    margin = 30
                    face_crop = frame[max(0, y - margin):y + bh + margin,
                    max(0, x - margin):x + bw + margin]

                    if face_crop.size != 0:
                        if img_count < TARGET_IMAGE_COUNT:
                            # Save the image
                            img_path = os.path.join(student_dir, f"{name}_{img_count}.jpg")
                            cv2.imwrite(img_path, face_crop)
                            img_count += 1
                            print(f"Captured: {img_count}/{TARGET_IMAGE_COUNT}")

                            # Display status on screen
                            cv2.putText(frame, f"Capturing: {img_count}/{TARGET_IMAGE_COUNT}",
                                        (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

            # Display the camera feed
            cv2.imshow('Face Data Recorder - Press Q to Quit', frame)

            k = cv2.waitKey(1) & 0xff
            # --- KAMERA ÇIKIŞ DÜZELTMESİ ---
            if k == ord('q'):
                break

            if img_count >= TARGET_IMAGE_COUNT:
                break
            # ---------------------------------

        print(f"\n✅ Data Capture Complete for: {student_name}. Total images: {img_count}")
        # Automatically take attendance upon successful data capture (Only once)
        if img_count > 0:
            save_attendance(student_name)

    # Kamera serbest bırakma ve pencereleri kapatma
    cam.release()
    cv2.destroyAllWindows()
    print("Camera closed.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        student_name = sys.argv[1]
        record_face_data(student_name)
    else:
        print("Usage: python face_data_recorder.py <Student_Name>")
        print("No student name provided. Exiting.")
